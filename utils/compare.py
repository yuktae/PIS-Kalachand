"""
Comparison Table Generator — aligns technical specifications across multiple
products into a single side-by-side table.

Public surface:
  align_specs(products)   -> dict {sections:[...], rows:[...], products:[...]}
  build_xlsx(table, out)  -> bytes (xlsx)
  build_csv(table)        -> str

The Gemini call clusters semantically equivalent spec keys; a deterministic
fallback runs after the call to catch obvious misses and to handle the case
where the model returns nothing usable.
"""
import io
import csv
import json
import re
from datetime import datetime
from typing import Any

from google.genai import types  # type: ignore[import-untyped]

from utils.ai_generation import _get_client, _require_prompt, _MODEL
from utils.json_utils import safe_json_loads


_MAX_PRODUCTS = 10
_SECTION_ORDER = [
    'General',
    'Dimensions & Weight',
    'Performance',
    'Features',
    'Energy & Environment',
    'Other',
]


def _collect_specs(product) -> dict[str, str]:
    """Pull the flat {label: value} spec dict off a Product. Empty dict if absent."""
    pis = product.pis_data or {}
    raw = pis.get('technical_specifications') or {}
    if not isinstance(raw, dict):
        return {}
    out: dict[str, str] = {}
    for k, v in raw.items():
        if not isinstance(k, str) or not k.strip():
            continue
        if v is None:
            continue
        out[k.strip()] = str(v).strip()
    return out


def _normalize_key(label: str) -> str:
    """Token-set normalisation for the deterministic fallback merge.
    'Power Consumption' / 'power-consumption' / 'POWER CONSUMPTION (W)' all collapse."""
    s = label.lower()
    s = re.sub(r'\([^)]*\)', '', s)            # drop parenthetical units
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    return s


def _build_table_payload(products, category_context: str) -> dict[str, Any]:
    """Shape the Gemini input. Keeps ids as strings throughout so the model's
    JSON keys round-trip cleanly."""
    items = []
    for p in products:
        items.append({
            'id': str(p.id),
            'name': p.model_name or f'Product #{p.id}',
            'specs': _collect_specs(p),
        })
    return {
        'category_context': category_context or 'mixed',
        'products_json': json.dumps(items, ensure_ascii=False),
    }


def _call_gemini(products, category_context: str) -> dict[str, Any] | None:
    """Single JSON-mode Gemini call. Returns the parsed dict or None on failure."""
    try:
        template = _require_prompt('compare_align_specs')
        prompt = template.format(**_build_table_payload(products, category_context))
        response = _get_client().models.generate_content(
            model=_MODEL,
            contents=prompt,
            config=types.GenerateContentConfig(response_mime_type='application/json'),
        )
        parsed = safe_json_loads(getattr(response, 'text', '') or '', fallback=None)
        if not isinstance(parsed, dict):
            return None
        rows = parsed.get('rows')
        if not isinstance(rows, list):
            return None
        return parsed
    except Exception as e:
        print(f'⚠️  compare_align_specs Gemini call failed: {e}')
        return None


def _deterministic_align(products) -> dict[str, Any]:
    """Fallback: case-insensitive label match across products. Used when the
    AI call fails entirely. Always produces a usable table — never raises."""
    clusters: dict[str, dict[str, Any]] = {}
    for p in products:
        pid = str(p.id)
        for label, value in _collect_specs(p).items():
            key = _normalize_key(label)
            if not key:
                continue
            cluster = clusters.setdefault(key, {
                'canonical_label': label,
                'section': 'Other',
                'unit_hint': '',
                'values': {},
            })
            cluster['values'][pid] = value
    rows = list(clusters.values())
    # Place rows with more filled values first so the table opens with real data.
    rows.sort(key=lambda r: -sum(1 for v in r['values'].values() if v))
    return {'rows': rows}


def _merge_fallback_into_ai(ai_result: dict[str, Any], products) -> dict[str, Any]:
    """If the AI missed a label that exists in the products but isn't in any
    AI row, append it as a deterministic row. Also collapses AI rows whose
    canonical labels normalise to the same key (defensive against the model
    emitting near-duplicates)."""
    ai_rows = ai_result.get('rows') or []

    # Collapse AI-emitted duplicates by canonical_label key.
    seen: dict[str, dict[str, Any]] = {}
    for row in ai_rows:
        if not isinstance(row, dict):
            continue
        label = (row.get('canonical_label') or '').strip()
        if not label:
            continue
        key = _normalize_key(label)
        if key in seen:
            # Merge values into the existing row, preferring non-null.
            for pid, val in (row.get('values') or {}).items():
                if val and not seen[key]['values'].get(pid):
                    seen[key]['values'][pid] = val
        else:
            seen[key] = {
                'canonical_label': label,
                'section': row.get('section') or 'Other',
                'unit_hint': row.get('unit_hint') or '',
                'values': {str(pid): v for pid, v in (row.get('values') or {}).items()},
            }

    # Now figure out which input labels never made it into the AI output and
    # tack them on so nothing is silently dropped.
    ai_covered_keys = set(seen.keys())
    leftover: dict[str, dict[str, Any]] = {}
    for p in products:
        pid = str(p.id)
        for label, value in _collect_specs(p).items():
            nkey = _normalize_key(label)
            if not nkey or nkey in ai_covered_keys:
                continue
            cluster = leftover.setdefault(nkey, {
                'canonical_label': label,
                'section': 'Other',
                'unit_hint': '',
                'values': {},
            })
            cluster['values'][pid] = value

    rows = list(seen.values()) + list(leftover.values())

    # Ensure every product id has an entry (null when missing) so the
    # frontend can render the column without per-cell existence checks.
    all_ids = [str(p.id) for p in products]
    for row in rows:
        for pid in all_ids:
            row['values'].setdefault(pid, None)
        # Coerce empty strings to None so the UI's "—" placeholder kicks in.
        for pid in all_ids:
            v = row['values'].get(pid)
            if isinstance(v, str) and not v.strip():
                row['values'][pid] = None

    return {'rows': rows}


def _section_sort_key(section: str) -> int:
    try:
        return _SECTION_ORDER.index(section)
    except ValueError:
        return len(_SECTION_ORDER)


def _order_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort by section priority, then by number of filled values descending."""
    def filled_count(row):
        return sum(1 for v in (row.get('values') or {}).values() if v)
    return sorted(rows, key=lambda r: (_section_sort_key(r.get('section') or 'Other'),
                                       -filled_count(r),
                                       (r.get('canonical_label') or '').lower()))


def align_specs(products, category_context: str = '') -> dict[str, Any]:
    """Top-level entry point used by the blueprint.

    Returns a dict shaped for the frontend:
        {
          "products": [{id, name, image_url}, ...],
          "rows":     [{canonical_label, section, unit_hint, values: {pid: str|None}}],
          "sections": ["General", "Performance", ...]   # sections actually present
        }
    """
    if not products:
        return {'products': [], 'rows': [], 'sections': []}
    if len(products) > _MAX_PRODUCTS:
        products = list(products)[:_MAX_PRODUCTS]

    ai = _call_gemini(products, category_context)
    if ai and ai.get('rows'):
        merged = _merge_fallback_into_ai(ai, products)
    else:
        merged = _deterministic_align(products)

    rows = _order_rows(merged.get('rows') or [])

    # Build the product header list with resolved image URLs. Done here (not
    # in the blueprint) so the export path can reuse the same table dict.
    # Use the blueprint's resolver so local paths get the /static/ prefix —
    # otherwise the browser tries to load them relative to /compare/* and 404s.
    from blueprints.compare import _resolve_image_url
    header = []
    for p in products:
        pis = p.pis_data or {}
        hi = pis.get('header_info') or {}
        header.append({
            'id': str(p.id),
            'name': p.model_name or f'Product #{p.id}',
            'brand': hi.get('brand') or '',
            'model_number': hi.get('model_number') or '',
            'image_url': _resolve_image_url(p.image_path),
        })

    sections_present: list[str] = []
    for r in rows:
        s = r.get('section') or 'Other'
        if s not in sections_present:
            sections_present.append(s)

    return {
        'products': header,
        'rows': rows,
        'sections': sections_present,
    }


# ───────────────────────── EXPORTS ─────────────────────────

def _cell_display(value) -> str:
    """How a missing value renders in the export (matches the UI's em-dash)."""
    if value is None:
        return '—'
    s = str(value).strip()
    return s if s else '—'


def build_csv(table: dict[str, Any]) -> str:
    """Write the comparison table to a CSV string. UTF-8 friendly.

    First row carries each product's image URL — CSV can't embed binary
    images, so the URL is the most useful representation. Excel renders it
    as plain text but the user can paste into a browser to verify, or
    copy into a HYPERLINK() formula. Empty string for products with no
    image so the column count stays consistent.
    """
    products = table.get('products') or []
    rows = table.get('rows') or []
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Row 0 — image URLs (only emitted if at least one product has an image).
    if any(p.get('image_url') for p in products):
        writer.writerow(['Image'] + [p.get('image_url') or '' for p in products])

    # Row 1 — product names.
    writer.writerow(['Specification'] + [p.get('name', '') for p in products])
    if any(p.get('brand') or p.get('model_number') for p in products):
        writer.writerow([''] + [
            ' / '.join(filter(None, [p.get('brand'), p.get('model_number')]))
            for p in products
        ])

    current_section = None
    for row in rows:
        section = row.get('section') or 'Other'
        if section != current_section:
            writer.writerow([f'── {section} ──'] + [''] * len(products))
            current_section = section
        label = row.get('canonical_label') or ''
        if row.get('unit_hint'):
            label = f"{label} ({row['unit_hint']})"
        writer.writerow([label] + [_cell_display(row['values'].get(p['id']))
                                   for p in products])
    return buf.getvalue()


def build_xlsx(table: dict[str, Any]) -> bytes:
    """Render the table as an .xlsx with image headers, frozen panes, banded
    rows, and section dividers. Returns raw bytes for send_file."""
    from openpyxl import Workbook  # type: ignore[import-untyped]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
    from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

    products = table.get('products') or []
    rows = table.get('rows') or []

    wb = Workbook()
    # `wb.active` is typed Optional but is always populated for a fresh
    # Workbook(). Narrow the type so the rest of the function stays clean.
    ws = wb.active
    assert ws is not None
    ws.title = 'Comparison'

    # ── Column widths
    ws.column_dimensions['A'].width = 32
    for i in range(len(products)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 26

    # ── Header rows (3 rows tall: image area + name + brand/model)
    header_fill = PatternFill('solid', fgColor='0F172A')   # slate-900
    header_font = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
    sub_font = Font(color='CBD5E1', size=10, name='Calibri')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # Row 1 — image space (tall row)
    ws.row_dimensions[1].height = 80
    ws.cell(row=1, column=1, value='').fill = header_fill
    # Row 2 — product name
    ws.cell(row=2, column=1, value='Specification')
    ws.cell(row=2, column=1).font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = left
    # Row 3 — brand / model line
    ws.cell(row=3, column=1, value='')
    ws.cell(row=3, column=1).fill = header_fill

    # Try embedding product images. Failures are silent — the export still
    # works without images on systems without Pillow/network/whatever.
    try:
        from openpyxl.drawing.image import Image as XLImage  # type: ignore[import-untyped]
        import urllib.request
        import tempfile, os
        for idx, p in enumerate(products):
            col_letter = get_column_letter(2 + idx)
            ws.cell(row=1, column=2 + idx, value='').fill = header_fill
            ws.cell(row=2, column=2 + idx, value=p.get('name', ''))
            ws.cell(row=2, column=2 + idx).font = header_font
            ws.cell(row=2, column=2 + idx).fill = header_fill
            ws.cell(row=2, column=2 + idx).alignment = center
            sub_label = ' / '.join(filter(None, [p.get('brand'), p.get('model_number')]))
            ws.cell(row=3, column=2 + idx, value=sub_label)
            ws.cell(row=3, column=2 + idx).font = sub_font
            ws.cell(row=3, column=2 + idx).fill = header_fill
            ws.cell(row=3, column=2 + idx).alignment = center

            url = p.get('image_url') or ''
            if not url:
                continue
            try:
                if url.startswith('http'):
                    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.img')
                    urllib.request.urlretrieve(url, tmp.name)
                    img_path = tmp.name
                else:
                    # Local static path. `_resolve_image_url` returns paths
                    # like '/static/uploads/foo.jpg' (with the Flask static
                    # prefix). Strip the leading '/static/' before joining
                    # so we don't end up looking in `static/static/uploads/`.
                    rel = re.sub(r'^/?static/', '', url)
                    img_path = os.path.join('static', rel.lstrip('/'))
                    if not os.path.exists(img_path):
                        continue
                img = XLImage(img_path)
                img.width = 90
                img.height = 90
                ws.add_image(img, f'{col_letter}1')
            except Exception:
                continue
    except Exception:
        # Header still rendered, just without images.
        pass

    # ── Body
    band_fill = PatternFill('solid', fgColor='F8FAFC')
    section_fill = PatternFill('solid', fgColor='E2E8F0')
    section_font = Font(bold=True, color='0F172A', size=11)
    thin = Side(style='thin', color='E2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    current_row = 4
    current_section = None
    band = False
    for row in rows:
        section = row.get('section') or 'Other'
        if section != current_section:
            ws.cell(row=current_row, column=1, value=section)
            ws.cell(row=current_row, column=1).font = section_font
            ws.cell(row=current_row, column=1).fill = section_fill
            for c in range(2, 2 + len(products)):
                ws.cell(row=current_row, column=c, value='').fill = section_fill
            current_row += 1
            current_section = section
            band = False

        label = row.get('canonical_label') or ''
        if row.get('unit_hint'):
            label = f"{label} ({row['unit_hint']})"

        fill = band_fill if band else None
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).alignment = left
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=10)
        ws.cell(row=current_row, column=1).border = border
        if fill:
            ws.cell(row=current_row, column=1).fill = fill
        for idx, p in enumerate(products):
            v = _cell_display(row['values'].get(p['id']))
            cell = ws.cell(row=current_row, column=2 + idx, value=v)
            cell.alignment = center
            cell.border = border
            if fill:
                cell.fill = fill
        current_row += 1
        band = not band

    ws.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(table: dict[str, Any]) -> bytes:
    """Render the comparison table as a landscape PDF using ReportLab's
    Platypus layout engine. Platypus handles page breaks automatically and
    repeats the header row + image strip at the top of every page so the
    output never has orphan rows or images stranded above headers.

    Layout:
      - Page: A4 landscape, 12pt margins
      - Image strip (90×72 px per product) at the top, repeated each page
      - Header rows: product name + brand/SKU (dark navy band)
      - Body: section divider rows + spec rows with alternating row bands
    """
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage,
    )
    from reportlab.platypus.flowables import KeepTogether
    import os
    import tempfile
    import urllib.request

    products = table.get('products') or []
    rows = table.get('rows') or []
    n_prod = max(1, len(products))

    # ── Page setup ──────────────────────────────────────────────────────
    page_size = landscape(A4)
    margin = 10 * mm
    avail_width = page_size[0] - 2 * margin

    # Column widths: 40% to the spec label, the rest split evenly among products.
    label_w = avail_width * 0.30
    product_w = (avail_width - label_w) / n_prod

    # ── Resolve images: download remote, locate local. Return path or None. ──
    img_temp_files: list[str] = []  # track for cleanup

    def _resolve_image_path(url: str) -> str | None:
        if not url:
            return None
        try:
            if url.startswith('http://') or url.startswith('https://'):
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.img')
                urllib.request.urlretrieve(url, tmp.name)
                img_temp_files.append(tmp.name)
                return tmp.name
            # Local static path — strip leading `/static/` before joining.
            rel = re.sub(r'^/?static/', '', url)
            local = os.path.join('static', rel.lstrip('/'))
            return local if os.path.exists(local) else None
        except Exception:
            return None

    # ── Build per-cell flowables ────────────────────────────────────────
    label_style = ParagraphStyle(
        'label', fontName='Helvetica-Bold', fontSize=8.5,
        textColor=colors.HexColor('#0f172a'), leading=10, alignment=0,  # left
    )
    cell_style = ParagraphStyle(
        'cell', fontName='Helvetica', fontSize=8.5,
        textColor=colors.HexColor('#1e293b'), leading=10, alignment=1,  # center
    )
    header_name_style = ParagraphStyle(
        'name', fontName='Helvetica-Bold', fontSize=10,
        textColor=colors.white, leading=12, alignment=1,
    )
    header_sub_style = ParagraphStyle(
        'sub', fontName='Helvetica', fontSize=8,
        textColor=colors.HexColor('#cbd5e1'), leading=10, alignment=1,
    )
    section_style = ParagraphStyle(
        'section', fontName='Helvetica-Bold', fontSize=9.5,
        textColor=colors.HexColor('#0f172a'), leading=12, alignment=0,
    )

    def _img_flowable(url: str):
        path = _resolve_image_path(url)
        if not path:
            # Empty paragraph keeps the cell height consistent.
            return Paragraph('—', cell_style)
        try:
            img = RLImage(path)
            # Fit inside 90 × 70pt, preserving aspect ratio.
            iw, ih = img.imageWidth, img.imageHeight
            max_w, max_h = product_w - 12, 70
            ratio = min(max_w / iw, max_h / ih)
            img.drawWidth = iw * ratio
            img.drawHeight = ih * ratio
            return img
        except Exception:
            return Paragraph('—', cell_style)

    # ── Assemble the table rows ────────────────────────────────────────
    data: list[list] = []

    # Row 0 — image strip. Annotate explicitly so the empty-string placeholder
    # doesn't narrow the row type to list[str] (we append Image/Paragraph below).
    img_row: list[Any] = ['']  # empty cell under the spec-label column
    for p in products:
        img_row.append(_img_flowable(p.get('image_url') or ''))
    data.append(img_row)

    # Row 1 — product names
    name_row = [Paragraph('Specification', header_name_style)]
    for p in products:
        name_row.append(Paragraph(_html_escape(p.get('name', '')), header_name_style))
    data.append(name_row)

    # Row 2 — brand / model number sub-line
    if any(p.get('brand') or p.get('model_number') for p in products):
        sub_row = [Paragraph('', header_sub_style)]
        for p in products:
            sub = ' / '.join(filter(None, [p.get('brand'), p.get('model_number')]))
            sub_row.append(Paragraph(_html_escape(sub), header_sub_style))
        data.append(sub_row)

    # Track which row indexes are section dividers (different styling later)
    section_rows: list[int] = []
    # Track header row count for `repeatRows` (image + names + brand/sku)
    header_row_count = len(data)

    current_section = None
    for row in rows:
        section = row.get('section') or 'Other'
        if section != current_section:
            section_rows.append(len(data))
            section_cells = [Paragraph(_html_escape(section.upper()), section_style)] + [''] * n_prod
            data.append(section_cells)
            current_section = section

        label_text = row.get('canonical_label') or ''
        if row.get('unit_hint'):
            label_text = f"{label_text} ({row['unit_hint']})"

        body_row = [Paragraph(_html_escape(label_text), label_style)]
        for p in products:
            v = _cell_display(row['values'].get(p['id']))
            body_row.append(Paragraph(_html_escape(v), cell_style))
        data.append(body_row)

    # ── Style the table ────────────────────────────────────────────────
    col_widths = [label_w] + [product_w] * n_prod

    tbl = Table(data, colWidths=col_widths, repeatRows=header_row_count)
    style = TableStyle([
        # Image-row aesthetics — light grey strip
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8fafc')),
        ('ROWHEIGHT', (0, 0), (-1, 0), 78),
        # Header band (names + brand line) — slate-900
        ('BACKGROUND', (0, 1), (-1, header_row_count - 1), colors.HexColor('#0f172a')),
        # Body grid
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#e2e8f0')),
        # Padding all around
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        # Vertical centering everywhere
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    # Section-row styling
    for sr in section_rows:
        style.add('BACKGROUND', (0, sr), (-1, sr), colors.HexColor('#e2e8f0'))
        style.add('SPAN', (0, sr), (-1, sr))
        style.add('LEFTPADDING', (0, sr), (-1, sr), 8)
    # Alternating row bands for body rows that aren't section dividers
    body_start = header_row_count
    body_band = False
    for r in range(body_start, len(data)):
        if r in section_rows:
            body_band = False
            continue
        if body_band:
            style.add('BACKGROUND', (0, r), (-1, r), colors.HexColor('#f8fafc'))
        body_band = not body_band

    tbl.setStyle(style)

    # ── Build the document ─────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin,
        title='Product Comparison',
    )
    title_style = ParagraphStyle(
        'title', fontName='Helvetica-Bold', fontSize=14,
        textColor=colors.HexColor('#0f172a'), leading=18,
    )
    sub_title_style = ParagraphStyle(
        'subtitle', fontName='Helvetica', fontSize=9,
        textColor=colors.HexColor('#64748b'), leading=11,
    )
    # `elements` holds a mix of Paragraph / Spacer / Table — all Flowable
    # subclasses, but the inferred literal type is too narrow. Annotate as
    # list[Any] so the SimpleDocTemplate.build() call type-checks.
    elements: list[Any] = [
        Paragraph('Product Comparison', title_style),
        Paragraph(f'{len(rows)} rows · {n_prod} product{"s" if n_prod != 1 else ""} · '
                  f'{datetime.now().strftime("%Y-%m-%d %H:%M")}', sub_title_style),
        Spacer(1, 6),
        tbl,
    ]
    doc.build(elements)

    # Clean up temp image files we downloaded.
    for p in img_temp_files:
        try: os.unlink(p)
        except OSError: pass

    return buf.getvalue()


def _html_escape(s) -> str:
    """Light escape for ReportLab paragraphs — only the chars ReportLab's
    mini-XML parser actually cares about. Accepts non-string input
    (None, ints, etc.) because the table dict can carry mixed types."""
    if s is None:
        return ''
    return (str(s).replace('&', '&amp;')
                  .replace('<', '&lt;')
                  .replace('>', '&gt;'))
