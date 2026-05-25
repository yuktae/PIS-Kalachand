"""
Bulk-import wizard — multi-product proforma flow (sibling to single_wizard).

Phase A delivers:
    • Session store keyed by triage_token (in-memory, 30 min TTL).
    • `triage_scan()` — one fast Gemini Flash call returning the document
      summary (density / has_images / origin / cluster_shape) plus a
      per-row item list with variant grouping. NOT the full extraction.

Subsequent phases will add:
    • cluster ops (merge / split / move),
    • full extraction with `batch_id`-tagged Product persistence,
    • lazy enrichment (image, content research, category),
    • commit / discard.

Pattern mirrors `utils/single_wizard.py` so reviewers can read both
together. Don't duplicate helpers across the two modules — anything used
by both should move to a shared `utils/_wizard_common.py` later.
"""

import os
import json
import time
import uuid
import threading

from google import genai
from google.genai import types

from .gemini_settings import gemini_http_options
from .prompt_manager import get_prompt


# ── Structured console logging ───────────────────────────────────────────────
# Used by the enrich loop and the image-allocation helpers so the bulk-job
# stdout is grep-able and visually scannable. Wire log (NDJSON) helpers
# below still use the old emoji-only format — they're consumed by the
# wizard UI, not the operator console.

_BAR_MAJOR = "═" * 64
_BAR_MINOR = "─" * 64


def _con_section(title: str) -> None:
    print(f"\n{_BAR_MAJOR}\n  {title}\n{_BAR_MAJOR}")


def _con_subsection(title: str) -> None:
    print(f"\n{_BAR_MINOR}\n  {title}\n{_BAR_MINOR}")


def _con_step(label: str, msg: str = '') -> None:
    """Print a `[label]` step header. `msg` is appended on the same line."""
    if msg:
        print(f"  [{label}] {msg}")
    else:
        print(f"  [{label}]")


def _con_info(msg: str) -> None:
    print(f"      · {msg}")


def _con_ok(msg: str) -> None:
    print(f"      ✓ {msg}")


def _con_warn(msg: str) -> None:
    print(f"      ⚠ {msg}")


# ── Gemini client ────────────────────────────────────────────────────────────
_MODEL = 'gemini-2.5-flash'

# Phase 3.0: thread-local — see ai_generation.py for rationale. The bulk
# worker pool calls _extract_variant_pis from multiple threads; a shared
# client closes its httpx transport after the first thread finishes.
import threading as _threading
_thread_local = _threading.local()


def _get_client():
    c = getattr(_thread_local, 'client', None)
    if c is None:
        c = genai.Client(
            api_key=os.getenv('GOOGLE_API_KEY'),
            http_options=gemini_http_options(),
        )
        _thread_local.client = c
    return c


# ── Session store (in-memory, 30 min TTL) ───────────────────────────────────
# TODO(multi-worker): same as single_wizard — needs Redis when we run >1 worker
# without sticky routing.
_SESSIONS: dict[str, dict] = {}
_LOCK = threading.Lock()
_TTL_SECONDS = 30 * 60


def _gc_locked() -> None:
    now = time.time()
    expired = [k for k, v in _SESSIONS.items() if v.get('_expires', 0) < now]
    for k in expired:
        _SESSIONS.pop(k, None)


def create_session(initial: dict | None = None) -> str:
    token = uuid.uuid4().hex
    payload = dict(initial or {})
    payload['_expires'] = time.time() + _TTL_SECONDS
    with _LOCK:
        _gc_locked()
        _SESSIONS[token] = payload
    return token


def get_session(token: str) -> dict | None:
    if not token:
        return None
    with _LOCK:
        sess = _SESSIONS.get(token)
        if not sess:
            return None
        sess['_expires'] = time.time() + _TTL_SECONDS
        return sess


def update_session(token: str, **fields) -> None:
    if not token:
        return
    with _LOCK:
        sess = _SESSIONS.get(token)
        if not sess:
            return
        sess.update(fields)
        sess['_expires'] = time.time() + _TTL_SECONDS


def drop_session(token: str) -> None:
    if not token:
        return
    with _LOCK:
        _SESSIONS.pop(token, None)


# ── Structured NDJSON logger (shared shape with single_wizard) ──────────────

_SEP = "─" * 56


def log_step(title: str) -> str:
    line = f"\n{_SEP}\n  {title.upper()}\n{_SEP}"
    print(line)
    return json.dumps({"log": {"type": "sep", "text": title}}) + "\n"


def log_info(msg: str) -> str:
    print(f"  · {msg}")
    return json.dumps({"log": {"type": "info", "text": msg}}) + "\n"


def log_ok(msg: str) -> str:
    print(f"  ✓ {msg}")
    return json.dumps({"log": {"type": "ok", "text": msg}}) + "\n"


def log_warn(msg: str) -> str:
    print(f"  ⚠ {msg}")
    return json.dumps({"log": {"type": "warn", "text": msg}}) + "\n"


def log_err(msg: str) -> str:
    print(f"  ✗ {msg}")
    return json.dumps({"log": {"type": "err", "text": msg}}) + "\n"


def log_progress(pct: int, msg: str | None = None) -> str:
    payload: dict = {"progress": pct}
    if msg:
        payload["message"] = msg
    return json.dumps(payload) + "\n"


def log_payload(**fields) -> str:
    return json.dumps(fields) + "\n"


# ── Triage scan ─────────────────────────────────────────────────────────────

_DEFAULT_TRIAGE = {
    "summary": {
        "item_count":    0,
        "density":       "minimal",
        "has_images":    "none",
        "origin":        "unknown",
        "cluster_shape": "single",
        "notes":         "",
    },
    "items": [],
}


def _normalize_origin_hint(origin_hint: str | None) -> str:
    """User checkbox: True → 'kalachand_internal', False/None → 'external_supplier'.
    Keep 'unknown' available for callers that don't want to pre-bias the model.
    """
    h = (origin_hint or "").strip().lower()
    if h in ('kalachand', 'kalachand_internal', 'internal', 'true', '1', 'yes'):
        return 'kalachand_internal'
    if h in ('external', 'external_supplier', 'supplier', 'false', '0', 'no'):
        return 'external_supplier'
    return 'unknown'


def _validate_triage(parsed: dict) -> dict:
    """Coerce arbitrary AI output into the expected shape so downstream code
    doesn't have to defend against missing keys / wrong types."""
    out = json.loads(json.dumps(_DEFAULT_TRIAGE))   # deep copy
    if not isinstance(parsed, dict):
        return out

    summary = parsed.get('summary') or {}
    if isinstance(summary, dict):
        for k, default in out['summary'].items():
            v = summary.get(k, default)
            if k == 'item_count':
                try:
                    v = int(v) if v is not None else 0
                except (TypeError, ValueError):
                    v = 0
            elif not isinstance(v, str):
                v = str(v) if v is not None else default
            out['summary'][k] = v

    items_in = parsed.get('items')
    items_out: list[dict] = []
    if isinstance(items_in, list):
        for idx, raw in enumerate(items_in):
            if not isinstance(raw, dict):
                continue
            # Coerce source_pages to a sorted, deduped list of non-negative
            # ints. Defaults to [0] when the model omits the field — a
            # single-page proforma is the most common case and we never want
            # an empty list (the slicer treats "no pages" as "use the whole
            # document", which defeats the purpose of variant-aware slicing).
            raw_pages = raw.get('source_pages')
            pages_clean: list[int] = []
            if isinstance(raw_pages, list):
                seen_pgs: set[int] = set()
                for p in raw_pages:
                    try:
                        pi = int(p)
                    except (TypeError, ValueError):
                        continue
                    if pi >= 0 and pi not in seen_pgs:
                        seen_pgs.add(pi)
                        pages_clean.append(pi)
            elif isinstance(raw_pages, (int, float)):
                pi = int(raw_pages)
                if pi >= 0:
                    pages_clean = [pi]
            if not pages_clean:
                pages_clean = [0]
            pages_clean.sort()

            entry = {
                "row_index":     int(raw.get('row_index', idx)) if str(raw.get('row_index', idx)).lstrip('-').isdigit() else idx,
                "name":          str(raw.get('name') or '').strip(),
                "brand":         str(raw.get('brand') or '').strip(),
                "model_number":  str(raw.get('model_number') or '').strip(),
                "price":         str(raw.get('price') or '').strip(),
                "category_hint": str(raw.get('category_hint') or '').strip(),
                "has_image":     bool(raw.get('has_image', False)),
                "variant_group": (str(raw['variant_group']).strip()
                                  if raw.get('variant_group') else None),
                "source_pages":  pages_clean,
            }
            items_out.append(entry)

    out['items'] = items_out
    out['summary']['item_count'] = len(items_out) or out['summary']['item_count']
    return out


def _build_feedback_section(feedback: str | None) -> str:
    """Format reviewer feedback into a prompt block. Returns empty string
    when feedback is missing/blank so the prompt template stays clean.

    The feedback is treated as AUTHORITATIVE — the language explicitly tells
    the model to prefer the human's instruction over the default clustering
    heuristics so global directives ("split everything", "merge rows 3–5")
    take effect instead of getting silently overridden by the rules above.
    """
    fb = (feedback or '').strip()
    if not fb:
        return ''
    return (
        "\n═════════════════ REVIEWER FEEDBACK (AUTHORITATIVE) ═════════════════\n"
        "A human reviewed the previous triage and left the note below. THIS\n"
        "FEEDBACK OVERRIDES the default clustering and variant-detection\n"
        "rules stated elsewhere in this prompt. Apply it LITERALLY:\n"
        "  • \"split all into separate PIS\" / \"each row is its own product\"\n"
        "    → set variant_group = null on EVERY item, set cluster_shape\n"
        "    to \"distinct\".\n"
        "  • \"rows 3–5 are variants of the same wardrobe\" / \"merge X and Y\"\n"
        "    → group those rows under one variant_group label.\n"
        "  • \"rename row 4 to X\" / \"the brand is wrong, it's actually Y\"\n"
        "    → apply the rename verbatim.\n"
        "  • \"this is sparse, not detailed\" / \"these have no images\"\n"
        "    → update the summary classification accordingly.\n"
        "Only fields the feedback does NOT mention may keep their previous\n"
        "values. Do not second-guess the human.\n\n"
        f"FEEDBACK NOTE:\n{fb}\n"
        "════════════════════════════════════════════════════════════════════"
    )


_XLSX_EXTENSIONS = ('.xlsx', '.xlsm')


def _is_xlsx_path(path: str) -> bool:
    """True for any spreadsheet format we render to text instead of
    handing to Gemini's Files API. Kept narrow on purpose — `.xls`
    (legacy binary) is NOT included; openpyxl can't read it and we
    surface a clear rejection at upload time rather than silently
    treating it as an unknown blob."""
    return bool(path) and path.lower().endswith(_XLSX_EXTENSIONS)


def _xlsx_to_text(xlsx_path: str) -> str:
    """Render a workbook into a structured text snapshot that we can
    pass to Gemini in place of an uploaded file.

    Each row is prefixed with `[Row N]` (1-indexed) so the triage prompt
    can refer back to specific rows when the wizard later matches
    extracted images to items. Multi-sheet workbooks are concatenated
    with `=== Sheet: X ===` dividers so the model sees the full doc.

    Cells are joined with ' | ' — a separator that's effectively never
    used inside proforma text — so Gemini can parse columns cleanly.
    Empty rows are skipped to keep the text under Gemini's context
    budget; this matters for big workbooks with formatting padding.

    Returns a best-effort string; never raises. If openpyxl can't load
    the file the wrapper returns a short error marker instead so the
    triage prompt fails loudly rather than silently dropping the doc.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        return f"(openpyxl not available: {e})"

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return f"(could not read workbook {os.path.basename(xlsx_path)}: {e})"

    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"=== Sheet: {sheet.title} ===")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if not any(c is not None and str(c).strip() != '' for c in row):
                continue
            cells = []
            for c in row:
                if c is None:
                    cells.append('')
                else:
                    cells.append(str(c).replace('\n', ' ').replace('|', '/').strip())
            line = ' | '.join(cells).rstrip(' |')
            lines.append(f"[Row {row_idx}] {line}")
        lines.append('')
    return '\n'.join(lines)


def _xlsx_extract_images(xlsx_path: str, out_dir: str,
                         prefix: str) -> list[dict]:
    """Extract embedded images from an XLSX, save each to `out_dir`, and
    return their anchor metadata.

    Returned shape per image:
        {
            'path':   absolute filesystem path of the saved image,
            'name':   basename of the saved file,
            'row':    0-indexed anchor row (top-left of the image),
            'col':    0-indexed anchor col,
            'sheet':  sheet title,
        }

    Anchor row is what lets us pair each image with the correct product
    line later — most supplier proformas place the photo on the same
    spreadsheet row as the item description. Cluster-level photos
    (logos in the header rows) sort to the top automatically.

    Best-effort: any image whose bytes we can't recover is skipped. The
    function never raises so a malformed picture in one row doesn't fail
    the whole triage.
    """
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        print(f"⚠ _xlsx_extract_images: load failed: {e}")
        return []

    results: list[dict] = []
    img_idx = 0
    for sheet in wb.worksheets:
        sheet_images = getattr(sheet, '_images', None) or []
        for img in sheet_images:
            # Anchor row/col. openpyxl normalises both OneCellAnchor and
            # TwoCellAnchor to expose `._from`; the field is sometimes
            # absent on legacy/exotic anchors so default to 0,0.
            row = col = 0
            anchor = getattr(img, 'anchor', None)
            cell_from = getattr(anchor, '_from', None) if anchor is not None else None
            if cell_from is not None:
                row = int(getattr(cell_from, 'row', 0) or 0)
                col = int(getattr(cell_from, 'col', 0) or 0)

            # Read bytes. openpyxl's reading API has shifted across
            # versions: `_data` was a callable on older builds, raw bytes
            # on others; `ref` shows up on some image objects too. Walk
            # the known options and stop at the first that yields bytes.
            data: bytes | None = None
            for attr in ('_data', 'ref', 'path'):
                v = getattr(img, attr, None)
                if v is None:
                    continue
                if callable(v):
                    try:
                        out = v()
                        if isinstance(out, (bytes, bytearray)):
                            data = bytes(out)
                            break
                    except Exception:
                        continue
                elif isinstance(v, (bytes, bytearray)):
                    data = bytes(v)
                    break
            if not data:
                continue

            # Sniff format from the leading bytes so we save with the
            # right extension. Saves us forcing a re-encode that could
            # quietly degrade quality.
            ext = '.png'
            if data[:3] == b'\xff\xd8\xff':
                ext = '.jpg'
            elif data[:4] == b'\x89PNG':
                ext = '.png'
            elif data[:6] in (b'GIF87a', b'GIF89a'):
                ext = '.gif'
            elif data[:4] == b'RIFF' and data[8:12] == b'WEBP':
                ext = '.webp'

            out_name = f'{prefix}_xlsx_img_{img_idx}{ext}'
            out_path = os.path.join(out_dir, out_name)
            try:
                with open(out_path, 'wb') as fh:
                    fh.write(data)
            except OSError as e:
                print(f"⚠ _xlsx_extract_images: save failed: {e}")
                continue

            results.append({
                'path':  out_path,
                'name':  out_name,
                'row':   row,
                'col':   col,
                'sheet': sheet.title,
            })
            img_idx += 1

    # Sort by (sheet, row, col) so callers can match items to images in
    # natural reading order — typically image[0] belongs to the row that
    # the first item lives on.
    results.sort(key=lambda r: (r['sheet'], r['row'], r['col']))
    return results


def xlsx_to_html_table(static_relpath: str, max_rows: int = 300) -> str:
    """Render a workbook as a sanitized HTML table for inline preview in
    the Source tab. `static_relpath` is a path under Flask's `static/`
    folder (e.g. 'uploads/PI_Foo.xlsx') — same convention used elsewhere
    in the audit-trail / proforma viewer.

    Layout strategy:
      • First pass collects every non-empty row, tracks which columns
        carry data anywhere in the sheet, and measures each column's
        widest cell so the second pass can size columns naturally.
      • Globally-empty columns (often Excel "spacer" columns) are
        dropped so the table doesn't waste horizontal real estate.
      • Merged cells round-trip as HTML `colspan` / `rowspan` so
        proforma header rows ("FOR & ON BEHALF OF BUYER" etc.) span
        their original width instead of squeezing into one narrow cell.
      • Long text is allowed to wrap (whitespace-pre-wrap) but each
        column gets enough room based on its widest entry, so short
        labels don't get word-wrapped just because they share a row
        with a paragraph.

    Output is a small chunk of HTML (one `<table>` per sheet) intended
    to live inside a scrollable container. Caps at `max_rows` per sheet
    so a 10k-row workbook can't tank the page. Never raises — returns
    an inline error block instead so the template keeps rendering.

    Exposed as a Jinja global named `render_xlsx_preview` so templates
    can write `{{ render_xlsx_preview(src) | safe }}` directly.
    """
    from flask import current_app
    from markupsafe import escape

    if not static_relpath:
        return '<p class="text-xs text-slate-400 italic p-4">No file.</p>'

    # Resolve under static/ — refuse anything escaping that root so
    # template input can't be coerced into reading arbitrary disk paths.
    static_root = os.path.join(current_app.root_path, 'static')
    abs_path = os.path.normpath(os.path.join(static_root, static_relpath))
    if not abs_path.startswith(static_root):
        return '<p class="text-xs text-red-500 p-4">Invalid file path.</p>'
    if not os.path.exists(abs_path):
        return f'<p class="text-xs text-red-500 p-4">File not found: {escape(static_relpath)}</p>'

    try:
        from openpyxl import load_workbook
    except Exception:
        return '<p class="text-xs text-red-500 p-4">openpyxl is required to preview Excel files.</p>'

    try:
        wb = load_workbook(abs_path, data_only=True)
    except Exception as e:
        return f'<p class="text-xs text-red-500 p-4">Cannot read workbook: {escape(str(e))}</p>'

    # Per-column width budget in pixels, derived from `len(content)`.
    # 6.5px/char fits the 11px font with breathing room. Floor/ceiling
    # keep columns from collapsing (numeric columns with short labels)
    # or running away (one giant paragraph cell that would force the
    # whole sheet wider than the Source pane).
    def _col_width_px(max_chars: int) -> int:
        return max(56, min(260, int(max_chars * 6.5) + 14))

    parts: list[str] = []
    for sheet in wb.worksheets:
        # ── Merged cells: map every covered position to its anchor and
        # record the colspan / rowspan to apply on that anchor cell.
        merge_anchor_of: dict[tuple[int, int], tuple[int, int]] = {}
        merge_span_of:   dict[tuple[int, int], tuple[int, int]] = {}
        for rng in sheet.merged_cells.ranges:
            top, left = rng.min_row, rng.min_col          # 1-indexed
            bot, right = rng.max_row, rng.max_col
            anchor = (top, left)
            merge_span_of[anchor] = (bot - top + 1, right - left + 1)
            for r in range(top, bot + 1):
                for c in range(left, right + 1):
                    merge_anchor_of[(r, c)] = anchor

        # ── First pass: collect rows + column stats ─────────────────
        # Keep absolute row index (1-indexed, openpyxl's native) so the
        # merge map can be consulted during render.
        collected: list[tuple[int, list[tuple[str, int]]]] = []  # (row_idx, [(text, col_idx_1based), ...])
        col_max_chars: dict[int, int] = {}
        row_seen = 0
        truncated = False
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_seen >= max_rows:
                truncated = True
                break
            if not any(c is not None and str(c).strip() != '' for c in row):
                continue
            row_seen += 1
            cells: list[tuple[str, int]] = []
            for offset, cell in enumerate(row):
                col_idx = offset + 1  # 1-indexed to match openpyxl conventions
                if cell is None:
                    text = ''
                else:
                    text = str(cell)
                cells.append((text, col_idx))
                if text.strip():
                    # Treat embedded newlines as separate "lines" — the
                    # column needs to be as wide as its longest LINE, not
                    # the whole multi-line cell, otherwise headers like
                    # "Buyer's\nItem No." would inflate the column.
                    line_width = max(len(seg) for seg in text.split('\n'))
                    col_max_chars[col_idx] = max(col_max_chars.get(col_idx, 0), line_width)
            collected.append((row_idx, cells))

        if not col_max_chars:
            # Sheet is entirely blank — skip the header too.
            continue

        # Visible column indexes, sorted left-to-right.
        visible_cols = sorted(col_max_chars.keys())

        # ── Render ─────────────────────────────────────────────────
        parts.append('<div class="mb-6">')
        # Sheet title scrolls with the content. The Source tab's
        # wrapper supplies its own sticky top bar (with the Download
        # button) so the user always has the action handy.
        parts.append(
            f'<h4 class="text-[10px] font-bold uppercase tracking-widest text-slate-500 mb-2 py-1.5 border-b border-slate-100">{escape(sheet.title)}</h4>'
        )
        parts.append(
            '<table class="text-[11px] border-collapse" style="border: 1px solid rgb(241 245 249); table-layout: fixed;">'
        )
        # <colgroup> with per-column widths.
        parts.append('<colgroup>')
        parts.append('<col style="width: 34px">')   # row-number gutter
        for col_idx in visible_cols:
            w = _col_width_px(col_max_chars.get(col_idx, 5))
            parts.append(f'<col style="width: {w}px">')
        parts.append('</colgroup>')

        # Track which (row, col) anchor cells we've already rendered so
        # the same merged content doesn't show up twice when its anchor
        # row hasn't reached the cells yet.
        rendered_anchors: set[tuple[int, int]] = set()

        display_row = 0
        for row_idx, cells in collected:
            display_row += 1
            parts.append(
                '<tr class="border-b border-slate-100 hover:bg-slate-50/60">'
            )
            parts.append(
                f'<td class="px-1 py-1 text-slate-300 text-[9px] align-top tabular-nums select-none border-r border-slate-100">{display_row}</td>'
            )
            # Map from col_idx → cell text for fast lookup.
            cell_text_by_col = {c_idx: text for text, c_idx in cells}

            skip_until_col: int | None = None  # used inside a horizontal merge
            for col_idx in visible_cols:
                # If this position is inside a merge and not the anchor,
                # the anchor handles it — render nothing for the slave.
                anchor = merge_anchor_of.get((row_idx, col_idx))
                if anchor is not None and anchor != (row_idx, col_idx):
                    continue

                attrs = ''
                if anchor == (row_idx, col_idx):
                    rowspan, colspan = merge_span_of[anchor]
                    if rowspan > 1:
                        attrs += f' rowspan="{rowspan}"'
                    if colspan > 1:
                        attrs += f' colspan="{colspan}"'

                text = cell_text_by_col.get(col_idx, '')
                # Preserve intentional newlines from the cell content;
                # let CSS handle wrap-by-space for very long lines.
                if text:
                    safe_text = escape(text).replace('\n', Markup_BR)
                else:
                    safe_text = ''
                parts.append(
                    f'<td{attrs} class="px-2 py-1 align-top text-slate-700 border-l border-slate-50" style="white-space: pre-wrap; word-break: break-word;">{safe_text}</td>'
                )
            parts.append('</tr>')

        parts.append('</table>')
        if truncated:
            remaining = max(0, sheet.max_row - row_seen)
            parts.append(
                f'<p class="text-[10px] italic text-slate-400 mt-2">'
                f'… preview truncated at {row_seen} rows (~{remaining} more in the sheet).'
                f'</p>'
            )
        parts.append('</div>')

    if not parts:
        return '<p class="text-xs text-slate-400 italic p-4">Empty workbook.</p>'
    return '\n'.join(parts)


# Pre-built Markup snippet for newline→<br> conversion inside escaped
# cell text. Stored as a constant so we don't import markupsafe at
# every cell render and don't risk the escape() output being re-escaped.
from markupsafe import Markup as _Markup
Markup_BR = _Markup('<br>')


def triage_scan(file_paths: list[str], origin_hint: str | None = None,
                feedback: str | None = None) -> dict:
    """One fast Gemini Flash call. Uploads the FIRST file, asks for the
    document summary + per-row preview list. Returns the validated JSON
    (always conformant to `_DEFAULT_TRIAGE` shape, never raises).

    `feedback`: optional free-text from the reviewer applied to a re-scan
    (Phase B "Re-run with feedback" button). Empty/None → first-run scan.

    NOTE: bulk imports often have a single proforma PDF spanning many pages
    plus optional supplementary spec sheets. For triage we only feed the
    first file — extraction (later phase) will see all of them.
    """
    if not file_paths:
        return _validate_triage({})

    fp = file_paths[0]
    if not fp or not os.path.exists(fp):
        return _validate_triage({})

    prompt_template = get_prompt('bulk_triage_scan') or ''
    prompt = prompt_template.format(
        origin_hint=_normalize_origin_hint(origin_hint),
        feedback_section=_build_feedback_section(feedback),
    )

    try:
        # XLSX path — Gemini's Files API doesn't ingest spreadsheets
        # reliably, and the tabular content is much cleaner as a row-
        # labelled text dump anyway. We send the workbook contents
        # inline in the prompt and skip the Files API upload entirely.
        if _is_xlsx_path(fp):
            xlsx_text = _xlsx_to_text(fp)
            full_prompt = (
                prompt
                + "\n\nDOCUMENT (Excel proforma — extracted as text):\n"
                + "Each line is prefixed with [Row N] giving its 1-indexed\n"
                + "spreadsheet row. Cell separator is ' | '.\n\n"
                + xlsx_text
            )
            contents = [full_prompt]
        else:
            uploaded = _get_client().files.upload(file=fp)
            # Wait up to ~15s for processing. The SDK marks state/name as
            # Optional; if either is missing we treat the upload as
            # already-ready (best effort).
            for _ in range(30):
                state = getattr(uploaded, 'state', None)
                state_name = getattr(state, 'name', None) if state is not None else None
                file_name = getattr(uploaded, 'name', None)
                if state_name != "PROCESSING" or not file_name:
                    break
                time.sleep(0.5)
                uploaded = _get_client().files.get(name=file_name)
            contents = [prompt, uploaded]

        from .api_metering import gemini_call
        response = gemini_call(
            prompt_id='bulk_triage_scan',
            model=_MODEL,
            client=_get_client(),
            contents=contents,
            config=types.GenerateContentConfig(response_mime_type="application/json"),
        )
        parsed = json.loads(response.text or "{}")
        return _validate_triage(parsed)

    except Exception as e:
        print(f"⚠ bulk triage_scan failed: {e}")
        # Return a minimal shape so the UI can still render something useful.
        out = _validate_triage({})
        out['summary']['notes'] = f"Triage failed: {e}"
        return out


# ── Cluster grouping (derived view of items) ────────────────────────────────

def build_stub_pis_from_cluster(cluster: dict, items: list[dict],
                                batch_id: str, origin_hint: str,
                                cluster_index: int = 0,
                                source_filenames: list[str] | None = None,
                                triage_summary: dict | None = None,
                                xlsx_image_pool: list[dict] | None = None) -> dict | None:
    """Phase C — produce a minimal pis_data dict for ONE cluster from the
    triage items it covers. Skipped items are excluded. Returns None when
    every item in the cluster is skipped.

    The output is intentionally STUB-ONLY (no specs, no sales arguments, no
    range_overview yet) — Phase D's lazy enrichment populates those fields
    on demand. This keeps "Generate PIS" instant and lets the user start
    reviewing immediately. The wizard's contract:
      • For SINGLETON clusters → one PIS shaped from the lone row.
      • For VARIANT clusters   → one PIS where header_info comes from the
        first non-skipped row, and variants[] lists every non-skipped row.

    `_bulk_*` keys are added so Phase D's workspace can re-load this draft
    by batch and remember which triage rows informed it.
    """
    active_indexes = [
        idx for idx in cluster.get('item_indexes', [])
        if 0 <= idx < len(items) and not items[idx].get('skip')
    ]
    if not active_indexes:
        return None

    primary = items[active_indexes[0]] or {}
    is_variant_cluster = (cluster.get('kind') == 'variants' and len(active_indexes) > 1)

    # For variant clusters the stub should already look like a family-level
    # PIS — the cluster label is the GENERAL name (e.g. "Sunon 4D Wardrobe")
    # and model_number aggregates every variant's SKU. The previous
    # implementation copied the FIRST item's specific name (e.g. "4D
    # WARDROBE-OAK/WARM WHITE") and only one SKU, which made the stub —
    # and the post-enrich state, since header_info wasn't being overwritten
    # — read like a single-product PIS.
    if is_variant_cluster:
        cluster_label = (cluster.get('label') or '').strip()
        # Concatenate every variant's SKU, deduping while preserving order.
        sku_seen, sku_list = set(), []
        for idx in active_indexes:
            sku = ((items[idx] or {}).get('model_number') or '').strip()
            if sku and sku not in sku_seen:
                sku_seen.add(sku)
                sku_list.append(sku)
        header = {
            'product_name':   cluster_label or (primary.get('name') or '').strip(),
            'brand':          (primary.get('brand') or '').strip(),
            'model_number':   ', '.join(sku_list),
            'price_estimate': (primary.get('price') or '').strip(),
        }
    else:
        header = {
            'product_name':  (primary.get('name') or '').strip(),
            'brand':         (primary.get('brand') or '').strip(),
            'model_number':  (primary.get('model_number') or '').strip(),
            'price_estimate': (primary.get('price') or '').strip(),
        }

    variants: list[dict] = []
    if is_variant_cluster:
        # Each row becomes a variant entry. The primary's variant label
        # falls out from `name` so the user can see it next to the rest.
        # `source_pages` is preserved per-variant so the variant-aware
        # image pipeline can slice the proforma into mini-PDFs (one per
        # variant) before doing extraction.
        for idx in active_indexes:
            it = items[idx] or {}
            variants.append({
                'label':         (it.get('name') or '').strip(),
                'model_number':  (it.get('model_number') or '').strip(),
                'price':         (it.get('price') or '').strip(),
                'source_pages':  list(it.get('source_pages') or [0]),
            })

    pis = {
        'header_info':            header,
        'range_overview':         '',
        'sales_arguments':        [],
        'technical_specifications': {},
        'warranty_service':       {'period': '', 'coverage': ''},
        'seo_data':               {'generated_keywords': '', 'meta_title': '',
                                   'meta_description': '', 'seo_long_description': ''},
        # Phase B carry-overs: variants stay attached to the cluster.
        'variants':               variants,

        # Phase C bookkeeping (so Phase D's workspace can find/reload).
        '_bulk_batch_id':         batch_id,
        '_bulk_cluster_index':    cluster_index,
        '_bulk_cluster_kind':     cluster.get('kind') or 'singleton',
        '_bulk_cluster_label':    cluster.get('label') or header['product_name'],
        '_bulk_row_indexes':      list(active_indexes),
        # Aggregated source_pages across every active row in the cluster —
        # used by the variant-aware image pipeline to produce a single mini-
        # PDF when the cluster spans multiple rows on different pages
        # (e.g. a wardrobe family with each finish printed on its own page).
        '_bulk_source_pages':     sorted({
            p for idx in active_indexes
            for p in (items[idx] or {}).get('source_pages') or [0]
            if isinstance(p, int) and p >= 0
        }) or [0],
        '_bulk_origin_hint':      origin_hint,
        # Phase D: filenames let the enricher rebuild absolute paths via
        # current_app.config['UPLOAD_FOLDER'] long after the wizard session
        # has expired. The files themselves stay on disk under uploads/.
        '_bulk_source_filenames': list(source_filenames or []),
        # Mirror onto the legacy `_source_files` key used by
        # verify_marketing.html's right-side Source tab (web-relative paths
        # under /static). This way bulk drafts get the same proforma viewer
        # the single-product wizard has had since Phase 2.5.
        '_source_files': [f"uploads/{fn}" for fn in (source_filenames or []) if fn],
        # Carry the triage summary so the enricher can route image/content
        # tasks correctly (sparse + has_images='none' → web search;
        # detailed + has_images='all' → doc-only crops; etc.).
        '_bulk_triage_density':    (triage_summary or {}).get('density', 'minimal'),
        '_bulk_triage_has_images': (triage_summary or {}).get('has_images', 'none'),
        '_enrichment_status':     'pending',
        '_enrichment_tasks':      {
            'image':    'pending',
            'content':  'pending',
            'category': 'pending',
        },
    }

    # Use the cluster label as the displayed model_name when the row
    # didn't print a clear name. Falls back to model_number, then "Item N+1".
    model_name = (header['product_name']
                  or pis['_bulk_cluster_label']
                  or header['model_number']
                  or f"Item {cluster_index + 1}")
    pis['_bulk_model_name'] = model_name

    # XLSX-embedded images, when present, slot directly into
    # `_bulk_image_candidates` so the Edit-PIS gallery shows them as
    # source-tagged candidates exactly like PDF-extracted ones.
    # Positional matching: items[] is in spreadsheet order, the image
    # pool is sorted by anchor row, so the Nth image of the workbook
    # is presumed to belong to the Nth item. If counts diverge we still
    # attach the available images at the equivalent positions and let
    # the user re-assign through the gallery — better than dropping
    # them silently.
    if xlsx_image_pool:
        cluster_candidates: list[dict] = []
        seen_paths: set[str] = set()
        for active_pos, item_idx in enumerate(active_indexes):
            # Prefer matching the global item_idx (covers the case where
            # earlier clusters didn't consume their slot — e.g. items
            # marked `skip`). Fall back to active_pos so a small workbook
            # with images on every active item still pairs correctly.
            picks = []
            if item_idx < len(xlsx_image_pool):
                picks.append(xlsx_image_pool[item_idx])
            if active_pos < len(xlsx_image_pool) and active_pos != item_idx:
                picks.append(xlsx_image_pool[active_pos])

            variant_sku = ''
            if is_variant_cluster and active_pos < len(variants):
                variant_sku = (variants[active_pos].get('model_number') or '').strip()

            for img in picks:
                p = img.get('path') or ''
                if not p or p in seen_paths:
                    continue
                seen_paths.add(p)
                cluster_candidates.append({
                    'path':        p,
                    'source':      'xlsx_embedded',
                    'variant_sku': variant_sku,
                })
        if cluster_candidates:
            pis['_bulk_image_candidates'] = cluster_candidates
            # Default the hero thumbnail to the first candidate so the
            # workspace card has a picture immediately, no enrichment
            # needed. Enrichment can still overwrite if it finds a
            # better web candidate later.
            pis['_image_path'] = cluster_candidates[0]['path']

    return pis


def _extract_variant_pis(file_paths: list[str], primary_name: str,
                         brand: str, variants: list[dict]) -> dict:
    """Run the bulk_variant_pis_extraction prompt against the uploaded
    document(s) — produces ONE PIS dict covering every variant in the
    cluster. Returns the validated dict (always conformant to the
    pis_extraction shape) or {} on failure so the caller can fall back."""
    if not file_paths or not primary_name or not variants:
        return {}

    # Format the variants block as a compact bulleted list the prompt can
    # quote back at the model.
    lines = []
    for v in variants:
        if not isinstance(v, dict):
            continue
        label = (v.get('label') or '').strip() or '(unnamed)'
        sku   = (v.get('model_number') or '').strip()
        price = (v.get('price') or '').strip()
        parts = [f"  - {label}"]
        if sku:   parts.append(f"· {sku}")
        if price: parts.append(f"· {price}")
        lines.append(' '.join(parts))
    variants_block = "\n".join(lines) if lines else "  (no variants listed)"

    prompt_template = get_prompt('bulk_variant_pis_extraction') or ''
    if not prompt_template:
        return {}
    prompt = prompt_template.format(
        primary_name=primary_name,
        brand=brand or '(unknown)',
        variants_block=variants_block,
        web_context='',
    )

    try:
        client = _get_client()
        # Upload every non-XLSX doc so the model sees full context;
        # XLSX files are rendered inline as text (see triage_scan for the
        # same rationale — Gemini's Files API doesn't ingest spreadsheets
        # reliably).
        uploaded = []
        xlsx_blocks: list[str] = []
        for fp in file_paths:
            if _is_xlsx_path(fp):
                xlsx_blocks.append(
                    f"--- {os.path.basename(fp)} (Excel) ---\n"
                    + _xlsx_to_text(fp)
                )
                continue
            uf = client.files.upload(file=fp)
            for _ in range(30):
                state = getattr(uf, 'state', None)
                state_name = getattr(state, 'name', None) if state is not None else None
                file_name = getattr(uf, 'name', None)
                if state_name != "PROCESSING" or not file_name:
                    break
                time.sleep(0.5)
                uf = client.files.get(name=file_name)
            uploaded.append(uf)

        full_prompt = prompt
        if xlsx_blocks:
            full_prompt += (
                "\n\nADDITIONAL EXCEL CONTENT (extracted as text, [Row N] = 1-indexed row):\n"
                + "\n\n".join(xlsx_blocks)
            )
        contents = [full_prompt] + uploaded
        from .api_metering import gemini_call
        response = gemini_call(
            prompt_id='bulk_variant_pis_extraction',
            model=_MODEL,
            client=client,
            contents=contents,
            config=types.GenerateContentConfig(response_mime_type="application/json"),
        )
        # Use the project's tolerant parser — Gemini occasionally returns
        # markdown-fenced JSON or a stray trailing comma. Plain json.loads
        # bails on those (we hit "Expecting ',' delimiter" on the 4D
        # wardrobe variant cluster); safe_json_loads cleans them up.
        from utils.json_utils import safe_json_loads
        parsed = safe_json_loads(response.text or "", fallback={})
        return parsed if isinstance(parsed, dict) else {}
    except Exception as e:
        print(f"⚠ _extract_variant_pis failed: {e}")
        return {}


def _resolve_source_paths(filenames: list[str], upload_folder: str) -> list[str]:
    """Turn `_bulk_source_filenames` back into absolute paths via the
    current upload folder. Filters out any that no longer exist on disk
    (the user may have manually cleaned the uploads dir)."""
    resolved: list[str] = []
    for fn in filenames or []:
        if not fn:
            continue
        p = os.path.join(upload_folder, fn)
        if os.path.exists(p):
            resolved.append(p)
    return resolved


def _derive_product_type_query(out: dict, brand: str) -> str:
    """Build a generic product-type search query for the case where the
    exact model-name search returns nothing — typical for SKU-style
    proforma rows like "00438 POB ARIETE CREAM/BLINT" that no SERP knows.

    Pulls product type from category_data (if the classifier has run),
    then from the AI-deduced range_overview, falling back to a brand-only
    query. Color/finish hints come from the first variant label when the
    cluster is variant-shaped.

    Example outputs:
      "Ariete electric kettle cream"   (brand + type + color)
      "Sunon wardrobe oak"             (brand + type + finish)
      "Ariete electric kettle"         (no color hint available)
      ""                               (no signal at all → caller skips)
    """
    parts: list[str] = []
    if brand:
        parts.append(brand)

    # Product type from the category classifier (most reliable signal).
    cat = out.get('category_data') or {}
    if isinstance(cat, dict):
        for key in ('product_type', 'subcategory', 'category'):
            v = (cat.get(key) or '').strip()
            if v:
                parts.append(v)
                break

    # Fallback: first noun-phrase of the range_overview narrative.
    if len(parts) < 2:
        narrative = (out.get('range_overview') or '').strip()
        if narrative:
            # Take the first 6 words, drop common filler.
            stop = {'the', 'a', 'an', 'this', 'these', 'with', 'for', 'of'}
            tokens = [w for w in narrative.split()[:6]
                      if w.lower() not in stop]
            if tokens:
                parts.append(' '.join(tokens[:3]))

    # Color/finish from the first variant label (helps narrow visual search).
    variants = out.get('variants') or []
    if isinstance(variants, list) and variants:
        v0 = variants[0] if isinstance(variants[0], dict) else {}
        label = (v0.get('label') or '').strip()
        if label and label.lower() not in ' '.join(parts).lower():
            parts.append(label)

    return ' '.join(parts).strip() if len(parts) >= 2 else ''


def enrich_product(pis_data: dict, upload_folder: str,
                   tasks: list[str] | None = None) -> dict:
    """Phase D — fill in the rich PIS fields for ONE bulk-import draft.

    Runs three enrichment tasks in order so the image search can lean on
    content + category data when it runs:
        content  → category  → image

    Each task updates its own slot in `_enrichment_tasks` ('done' |
    'failed' | 'skipped') so the workspace UI can render per-task status.

    `tasks` filters which jobs to run (defaults to all three). The caller
    persists the returned dict back to `Product.pis_data`. The function is
    idempotent — running it twice on the same draft re-runs the enrichers
    and overwrites their last result, but never touches user-edited fields
    in `header_info` (those are treated as authoritative).

    Returns the updated pis_data dict (NOT mutated in place — caller can
    diff cleanly).
    """
    import copy

    out = copy.deepcopy(pis_data or {})
    out.setdefault('_enrichment_tasks', {'image': 'pending',
                                          'content': 'pending',
                                          'category': 'pending'})
    wanted = set(tasks or ['image', 'content', 'category'])

    file_paths = _resolve_source_paths(out.get('_bulk_source_filenames') or [],
                                       upload_folder)
    header = out.get('header_info') or {}
    target_name = (header.get('product_name') or '').strip() \
                   or (out.get('_bulk_cluster_label') or '').strip()
    brand = (header.get('brand') or '').strip()
    has_images = (out.get('_bulk_triage_has_images') or 'none').lower()
    # Variants. Each variant has its own SKU + label; the SKU is what gets
    # attached to image candidates as `variant_sku` so the Edit-PIS gallery
    # can show source badges and route Set-as-main / Assign actions
    # correctly. The cluster primary uses key='__primary__' as its handle.
    variant_meta_for_alloc: list[dict] = []
    variant_names: list[str] = []
    for v in (out.get('variants') or []):
        if not isinstance(v, dict):
            continue
        label = (v.get('label') or '').strip()
        sku   = (v.get('model_number') or '').strip()
        if label and label.lower() != target_name.lower():
            variant_names.append(label)
        if sku:
            variant_meta_for_alloc.append({'key': sku, 'label': label or sku})

    # Per-cluster source pages (zero-based) — narrows the embedded scan to
    # just the pages this cluster spans. Triage gives us one set per
    # cluster; per-variant page indexes live inside variants[*].source_pages
    # but we don't need them at allocation time (one Gemini call sees them
    # all).
    cluster_pages: list[int] = list(out.get('_bulk_source_pages') or [])

    # ── Console banner for this cluster ───────────────────────────────────
    cluster_kind = (out.get('_bulk_cluster_kind') or 'singleton').lower()
    variant_count = len(out.get('variants') or [])
    if cluster_kind == 'variants' and variant_count > 1:
        kind_blurb = f"variant cluster · {variant_count} variants"
    else:
        kind_blurb = 'singleton'
    _con_section(f"CLUSTER — {target_name or 'unnamed'}  ({kind_blurb})")
    _con_info(f"has_images: {has_images}  ·  pages: {cluster_pages or 'all'}")

    # ── Task: image — extracts for primary AND every variant ──────────
    # Runs LAST (after content + category) so the product-type fallback
    # query can use category_data when the SKU-style model name fails.
    #
    # Routing by triage signal:
    #   has_images in ('all', 'partial')  — doc-side bbox/embedded
    #     extraction is the primary source; web is a backup.
    #   has_images == 'none'              — text-only proforma. NO doc
    #     extraction, NO auto-AI. Two-tier web only:
    #       Tier 1: supplier URL discovery + scrape via the discovered
    #               URL (same multi-engine cascade single uses).
    #       Tier 2: product-type search ("Brand + category + color"),
    #               only fires when Tier 1 returns nothing.
    #     Capped at 2 candidates per variant.
    def _image_task() -> dict:
        """Two-branch image pipeline. Drops every method that's already
        available from the Edit-PIS gallery (nano-banana isolate, dead-site
        scrapes, product-type fallback) — those are reachable on demand
        via the per-product `/api/product/<id>/image/*` endpoints.

        Branch A — proforma has photos (`has_images` in 'all', 'partial'):
          1. Per-variant PDF/embedded extract. The triage gave us a per-row
             `source_pages` allocation; `extract_specific_image` uses the
             row text neighbourhood to pick the right photo per variant.
          2. One cluster-level smart web search (max 2 candidates).
          STOP.

        Branch B — proforma has no photos (`has_images` == 'none'):
          1. One cluster-level smart web search.
          2. Discover supplier URL and scrape it for images.
          STOP.

        Everything else (nano-banana, supplier-page scrape for Branch A,
        product-type fallback search) is now opt-in from Edit PIS.
        """
        # Late imports — image_processing has heavy deps (Playwright,
        # PIL, etc.) so we don't pay for them when only content is enriched.
        from utils.single_wizard import extract_image_candidates_from_web
        from concurrent.futures import ThreadPoolExecutor
        from .pdf_processing import (
            extract_and_allocate_embedded, extract_specific_image,
            extract_product_from_image,
        )
        result = {'image_path': None, 'image_candidates': []}
        seen_paths: set[str] = set()

        def _push(path: str | None, source: str, page_url: str | None,
                  variant_sku: str | None) -> None:
            """Append a candidate to the list. `variant_sku` is the SKU we
            tag the candidate with (so the Edit-PIS gallery can route
            assign / set-as-main correctly). None = cluster-level."""
            if not path or path in seen_paths:
                return
            seen_paths.add(path)
            entry: dict = {'path': path, 'page_url': page_url, 'source': source}
            if variant_sku:
                entry['variant_sku'] = variant_sku
            result['image_candidates'].append(entry)

        try:
            if has_images in ('all', 'partial'):
                # ══ Branch A — PDF has embedded photos ═══════════════════
                # PDF embedded extraction (CPU + Gemini Vision) and the
                # web search (network + Gemini grounding) are independent,
                # so we kick both off in parallel and join below. Saves
                # roughly max(pdf, web) - min(pdf, web) per cluster.
                _con_step('IMAGES', 'branch=pdf-embedded')
                fp = file_paths[0] if file_paths else None

                def _do_pdf_extract():
                    if fp and os.path.splitext(fp)[1].lower() == '.pdf':
                        try:
                            return extract_and_allocate_embedded(
                                fp,
                                variant_meta_for_alloc or [{'key': '__primary__',
                                                            'label': target_name or 'product'}],
                                upload_folder,
                                page_filter=cluster_pages or None,
                                cluster_label=target_name or '',
                            ) or {'images': [], 'allocations': {}, 'unallocated': []}
                        except Exception as e:
                            _con_warn(f"embedded extract+allocate failed: {e}")
                            return {'images': [], 'allocations': {}, 'unallocated': []}
                    if fp:
                        # Image proforma (no PDF) — bbox extract per variant.
                        bbox_paths: list[str] = []
                        for tgt in ([target_name] + variant_names):
                            if not tgt:
                                continue
                            try:
                                paths = extract_product_from_image(
                                    fp, tgt, upload_folder,
                                    skip_verify=True, all_matches=True,
                                ) or []
                            except Exception as e:
                                _con_warn(f"image bbox for '{tgt}' failed: {e}")
                                paths = []
                            bbox_paths.extend(paths)
                        return {'images': [], 'allocations': {},
                                'unallocated': [], '_bbox_paths': bbox_paths}
                    return {'images': [], 'allocations': {}, 'unallocated': []}

                def _do_web_search():
                    if not target_name:
                        return []
                    try:
                        return extract_image_candidates_from_web(
                            model_name=target_name, supplier_url=None,
                            upload_folder=upload_folder, brand=brand or None,
                            max_results=2, log_cb=None,
                        ) or []
                    except Exception as e:
                        _con_warn(f"web search failed: {e}")
                        return []

                with ThreadPoolExecutor(max_workers=2) as _ex:
                    pdf_future = _ex.submit(_do_pdf_extract)
                    web_future = _ex.submit(_do_web_search)
                    alloc_result = pdf_future.result()
                    web = web_future.result()

                # Bbox-fallback paths from the standalone-image branch
                # are unallocated, cluster-level candidates.
                for p in (alloc_result.get('_bbox_paths') or []):
                    _push(p, 'document', None, None)

                # Step 2 — push allocations into the candidate list and
                # mirror onto variants[*].image_paths for the variant strip.
                if alloc_result.get('images'):
                    _con_info(f"{len(alloc_result['images'])} embedded photo(s) saved (one pass)")
                    _alloc_raw = alloc_result.get('allocations')
                    allocations: dict[str, list] = _alloc_raw if isinstance(_alloc_raw, dict) else {}
                    _unalloc_raw = alloc_result.get('unallocated')
                    unalloc: list = _unalloc_raw if isinstance(_unalloc_raw, list) else []
                    variants_list = out.get('variants') or []

                    for vkey, paths in allocations.items():
                        # Resolve label for log readability.
                        v_label = next(
                            (v.get('label') or v.get('model_number') or vkey
                             for v in variants_list
                             if isinstance(v, dict) and (v.get('model_number') or '').strip() == vkey),
                            vkey,
                        )
                        _con_ok(f"{vkey}  ← {len(paths)} photo(s)  [{v_label}]")
                        for p in paths:
                            _push(p, 'document', None, vkey)
                        # Mirror onto the variant's image_paths so the
                        # workspace's per-variant strip lights up.
                        for v in variants_list:
                            if not isinstance(v, dict):
                                continue
                            if (v.get('model_number') or '').strip() != vkey:
                                continue
                            existing = list(v.get('image_paths') or [])
                            for p in paths:
                                if p and p not in existing:
                                    existing.append(p)
                            v['image_paths'] = existing
                            if existing and not v.get('image_path'):
                                v['image_path'] = existing[0]
                            break

                    if unalloc:
                        _con_info(f"{len(unalloc)} photo(s) unallocated → cluster gallery")
                        for p in unalloc:
                            _push(p, 'document', None, None)

                # Step 3 — push the web candidates from the parallel search above.
                if web:
                    _con_info(f"{len(web)} web candidate(s)")
                    for r in web:
                        _push(r.get('path'), 'web', r.get('page_url'), None)
            else:
                # ══ Branch B — PDF has no embedded photos ════════════════
                # Single web search via Gemini grounding → top 2 URLs →
                # parallel page scrape → up to 2 image candidates total.
                # The legacy SUPPLIER auto-pass was redundant (it discovered
                # and re-scraped the same domains the WEB pass already hit)
                # and is now an Edit-PIS-only manual action.
                if target_name:
                    _con_step('IMAGES', 'web-only')
                    try:
                        web = extract_image_candidates_from_web(
                            model_name=target_name, supplier_url=None,
                            upload_folder=upload_folder, brand=brand or None,
                            max_results=2, log_cb=None,
                        ) or []
                    except Exception as e:
                        _con_warn(f"web search failed: {e}")
                        web = []
                    _con_info(f"{len(web)} web candidate(s)")
                    for r in web:
                        _push(r.get('path'), 'web', r.get('page_url'), None)

            # Default thumbnail — first 'document' candidate (Branch A) or
            # whatever the web/supplier search returned first (Branch B).
            doc_first = next(
                (c for c in result['image_candidates'] if c['source'] == 'document'),
                None,
            )
            if doc_first:
                result['image_path'] = doc_first['path']
            elif result['image_candidates']:
                result['image_path'] = result['image_candidates'][0]['path']
        except Exception as e:
            result['_error'] = f"Image task failed: {e}"
        return result

    # ── Task 2: content (range_overview, sales_arguments, specs, SEO, warranty) ──
    # Routes by cluster kind:
    #   singleton → existing pis_extraction prompt (same as single-mode wizard)
    #   variants  → new bulk_variant_pis_extraction prompt that produces ONE
    #               PIS covering ALL variants (mentions every variant in
    #               description, lists common specs with per-variant notes,
    #               highlights the range in sales arguments).
    def _content_task() -> dict:
        result: dict = {}
        try:
            if not file_paths or not target_name:
                result['_error'] = "Content task skipped (no source files or name)."
                return result

            cluster_kind = (out.get('_bulk_cluster_kind') or 'singleton').lower()
            variants_full = out.get('variants') or []

            from utils.ai_generation import generate_pis_data
            from utils.image_processing import gather_web_context_for_content

            # Phase 3.3: pull live product-page text from the same Brave-
            # discovered URLs the image pipeline uses, and feed it to
            # `generate_pis_data` as `web_context`. Without this, sparse
            # proformas (the Xiaomi case) yielded zero technical specs.
            # `_field_origin` grep-verifies each field against the proforma
            # raw text downstream — anything that came from web instead of
            # the proforma will continue to flag as AI-generated, so the
            # verify-PIS badges stay correct without further changes.
            web_context = ""
            if target_name:
                try:
                    web_context = gather_web_context_for_content(
                        target_name, brand=brand or None,
                        log_cb=lambda m: _con_info(f"web ctx: {m}"),
                    )
                except Exception as e:
                    _con_warn(f"web context fetch failed: {e}")
                    web_context = ""
            url_data = {"text": web_context, "html": ""}
            # Persist the exact Brave web text the generator saw so the
            # downstream origin classifier can split web_grounded from
            # hallucinated, and the AI-accuracy eval can re-use it as
            # the second source for LLM judging.
            if web_context:
                result['_web_context'] = web_context

            ai: dict = {}
            if cluster_kind == 'variants' and len(variants_full) > 1:
                ai = _extract_variant_pis(
                    file_paths, target_name, brand, variants_full,
                ) or {}
                # Variant extraction can return {} if Gemini's JSON had a
                # syntax error or the prompt template misfired. Fall back to
                # the same single-product extractor singleton clusters use,
                # so the user at least gets the primary variant's PIS rather
                # than an empty card.
                if not ai.get('range_overview'):
                    print("  ↩ Variant extraction empty — falling back to singleton path with primary name.")
                    ai = generate_pis_data(file_paths, target_name, url_data) or {}
            else:
                ai = generate_pis_data(file_paths, target_name, url_data) or {}

            if not isinstance(ai, dict):
                result['_error'] = "AI returned non-dict content"
                return result
            if not ai:
                result['_error'] = "AI returned empty content"
                return result
            # Copy the rich enrichment fields.
            for key in ('range_overview', 'sales_arguments',
                        'technical_specifications', 'warranty_service',
                        'seo_data'):
                if key in ai:
                    result[key] = ai[key]
            # For VARIANT clusters specifically, also surface the AI's
            # header_info — the variant prompt produces the right family
            # name + comma-separated SKUs that no stub heuristic can match.
            # The merger above will only apply this when the user hasn't
            # manually edited header_info (see `_user_edited_header` flag
            # set by the save endpoint).
            if cluster_kind == 'variants' and len(variants_full) > 1:
                ai_header = ai.get('header_info') or {}
                if isinstance(ai_header, dict) and ai_header:
                    result['_ai_header_info'] = ai_header
            # `seo_keywords` (Product column) is computed from seo_data later.
            seo = ai.get('seo_data') or {}
            if seo.get('generated_keywords'):
                result['_seo_keywords'] = seo['generated_keywords']
        except Exception as e:
            result['_error'] = f"Content task failed: {e}"
        return result

    # Sequential pass — content → category → image. Image runs LAST so
    # the product-type fallback query has access to category_data. The
    # previous parallel (image + content) layout meant the image task
    # always ran with empty content, which made the product-type tier
    # impossible.
    if 'content' in wanted:
        _con_step('CONTENT', 'AI extraction')
        r = _content_task() or {}
        if r.get('_error'):
            out['_enrichment_tasks']['content'] = 'failed'
            out.setdefault('_enrichment_errors', {})['content'] = r['_error']
            _con_warn(f"content failed: {r['_error']}")
        else:
            for key in ('range_overview', 'sales_arguments',
                        'technical_specifications', 'warranty_service',
                        'seo_data'):
                if key in r:
                    out[key] = r[key]
            if r.get('_web_context'):
                out['_web_context'] = r['_web_context']
            if r.get('_seo_keywords'):
                out['_seo_keywords_pending'] = r['_seo_keywords']
            # Apply AI header_info for variant clusters — but only when
            # the user hasn't manually overridden the header (the save
            # endpoint sets `_user_edited_header` once anything in
            # header_info is touched). On first enrich the flag isn't
            # set, so the family name + concatenated SKUs that the
            # variant prompt produces win out over the stub.
            ai_hdr = r.get('_ai_header_info') or {}
            if ai_hdr and not out.get('_user_edited_header'):
                cur = out.get('header_info') or {}
                merged = dict(cur)
                for hk in ('product_name', 'model_number',
                           'brand', 'price_estimate'):
                    v = ai_hdr.get(hk)
                    if isinstance(v, str) and v.strip():
                        merged[hk] = v.strip()
                out['header_info'] = merged

            # Origin map for the verify-PIS badge UI. Bulk extraction
            # doesn't split source_facts / ai_enriched_details the way
            # the single-product proforma flow does, so we grep-verify
            # each filled-in field against the uploaded Proforma's raw
            # text. Strict-fact rule applies: only Proforma-confirmed
            # values become 'verified' (yellow ✔); everything else
            # lands in the AI bucket (red ✨).
            try:
                from helpers import (
                    extract_raw_text_from_files,
                    classify_flat_pis_origins,
                )
                raw_doc_text = extract_raw_text_from_files(file_paths) or ""
                field_origins, spec_origins = classify_flat_pis_origins(
                    out, raw_doc_text,
                    web_context=out.get('_web_context', ''),
                )
                out['_field_origins'] = field_origins
                out['_spec_origins'] = spec_origins
            except Exception as e:
                print(f"⚠ origin classification failed for bulk PIS: {e}")

            out['_enrichment_tasks']['content'] = 'done'
            _con_ok("content enriched")

    # ── Category (depends on content being filled in) ──────────────────
    if 'category' in wanted:
        _con_step('CATEGORY', 'AI classification')
        try:
            from utils.category_classifier import classify_product_category
            result = classify_product_category(out) or {}
            if result and not result.get('error'):
                out['category_data'] = result
                out['_enrichment_tasks']['category'] = 'done'
                c1 = result.get('category_1', '')
                c2 = result.get('category_2', '')
                c3 = result.get('category_3', '')
                _con_ok(f"{c1} > {c2} > {c3}")
            else:
                out['_enrichment_tasks']['category'] = 'failed'
                _con_warn("category classification returned no result")
        except Exception as e:
            out['_enrichment_tasks']['category'] = 'failed'
            out.setdefault('_enrichment_errors', {})['category'] = f"Category task failed: {e}"
            _con_warn(f"category failed: {e}")

    # ── Image (runs LAST — needs content + category context) ───────────
    if 'image' in wanted:
        r = _image_task() or {}
        if r.get('_error'):
            out['_enrichment_tasks']['image'] = 'failed'
            out.setdefault('_enrichment_errors', {})['image'] = r['_error']
        else:
            if r.get('image_path'):
                out['_image_path'] = r['image_path']
            if r.get('image_candidates'):
                out['_bulk_image_candidates'] = r['image_candidates']
            out['_enrichment_tasks']['image'] = 'done'

    # Mark overall status. 'done' if every wanted task finished cleanly,
    # 'partial' when some failed, 'failed' when all wanted tasks failed.
    statuses = [out['_enrichment_tasks'].get(t, 'pending') for t in wanted]
    if all(s == 'done' for s in statuses):
        out['_enrichment_status'] = 'done'
    elif any(s == 'done' for s in statuses):
        out['_enrichment_status'] = 'partial'
    else:
        out['_enrichment_status'] = 'failed'

    return out


def derive_cluster_groups(items: list[dict]) -> list[dict]:
    """Bucket items by `variant_group`. Items with `variant_group=None` each
    get their own singleton group. Output is ordered: variant groups first
    (by first-appearance), then singletons (by row_index).

    Each group: {"id", "label", "kind": "variants"|"singleton", "item_indexes": [int]}.
    Indexes refer to the position in the input list (so the frontend can
    splice items by index when the user reshapes clusters).
    """
    groups: list[dict] = []
    seen: dict[str, int] = {}      # variant_group label → groups[] index
    for i, item in enumerate(items):
        vg = item.get('variant_group')
        if vg:
            if vg not in seen:
                seen[vg] = len(groups)
                groups.append({
                    'id':           f"g_{i}",
                    'label':        vg,
                    'kind':         'variants',
                    'item_indexes': [i],
                })
            else:
                groups[seen[vg]]['item_indexes'].append(i)
        else:
            groups.append({
                'id':           f"g_{i}",
                'label':        item.get('name') or f"Item {i+1}",
                'kind':         'singleton',
                'item_indexes': [i],
            })
    return groups
